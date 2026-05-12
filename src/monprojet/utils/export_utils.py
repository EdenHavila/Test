from csv import writer as csv_writer

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_csv_response(filename, headers, rows):
    """Construit une réponse CSV téléchargeable."""
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write("\ufeff")

    csv_file = csv_writer(response, delimiter=";")
    csv_file.writerow(headers)
    for row in rows:
        csv_file.writerow(row)

    return response


def build_excel_response(filename, sheet_title, headers, rows):
    """Construit un fichier Excel simple et lisible."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        worksheet.append(row)

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


